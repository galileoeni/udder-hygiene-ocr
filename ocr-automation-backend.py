import os
import re
import pandas as pd
import numpy as np
from datetime import datetime
from pathlib import Path
import pytesseract
from PIL import Image
import cv2
import fitz  # PyMuPDF for PDF handling
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.chart import BarChart, Reference
import json
from flask import Flask, request, jsonify, send_file
from werkzeug.utils import secure_filename
import logging
import io
import platform

# Configure Tesseract path for Windows
if platform.system() == 'Windows':
    # Update this path if Tesseract is installed elsewhere
    pytesseract.pytesseract.tesseract_cmd = r'C:\Program Files\Tesseract-OCR\tesseract.exe'

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

class UdderHygieneOCR:
    """Main OCR processing class for udder hygiene documents"""
    
    def __init__(self):
        self.supported_formats = ['.pdf', '.jpg', '.jpeg', '.png', '.tiff']
        self.data_pattern = re.compile(r'(\d{1,3})\s*(?:,|\s)\s*(\d{1,3})\s*(?:,|\s)\s*(\d{1,3})')
        self.group_pattern = re.compile(r'Group\s*([A-Z])', re.IGNORECASE)
        self.date_pattern = re.compile(r'(\d{1,2})[/-](\d{1,2})[/-](\d{2,4})')
        
    def preprocess_image(self, image_path):
        """Preprocess image for better OCR accuracy"""
        # Read image
        img = cv2.imread(str(image_path))
        
        # Convert to grayscale
        gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
        
        # Apply thresholding to get better OCR results
        thresh = cv2.threshold(gray, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)[1]
        
        # Denoise
        denoised = cv2.fastNlMeansDenoising(thresh)
        
        # Resize if too small
        height, width = denoised.shape
        if width < 1000:
            scale_factor = 1000 / width
            new_width = int(width * scale_factor)
            new_height = int(height * scale_factor)
            denoised = cv2.resize(denoised, (new_width, new_height))
        
        return denoised
    
    def pdf_to_images(self, pdf_path):
        """Convert PDF pages to images"""
        doc = fitz.open(pdf_path)
        images = []
        
        for page_num in range(len(doc)):
            page = doc[page_num]
            pix = page.get_pixmap(matrix=fitz.Matrix(300/72, 300/72))  # 300 DPI
            img_data = pix.pil_tobytes(format="PNG")
            images.append(Image.open(io.BytesIO(img_data)))
        
        doc.close()
        return images
    
    def extract_text_from_image(self, image):
        """Extract text from image using OCR"""
        # Configure tesseract for better results
        custom_config = r'--oem 3 --psm 6'
        
        try:
            text = pytesseract.image_to_string(image, config=custom_config)
            return text
        except Exception as e:
            logger.error(f"OCR error: {str(e)}")
            return ""
    
    def parse_ocr_text(self, text, filename=""):
        """Parse OCR text to extract udder hygiene data"""
        lines = text.split('\n')
        extracted_data = []
        
        # Extract farm name from filename or text
        farm_name = self.extract_farm_name(filename, text)
        
        # Extract date
        date = self.extract_date(text)
        
        current_group = None
        
        for line in lines:
            # Check for group identifier
            group_match = self.group_pattern.search(line)
            if group_match:
                current_group = f"Group {group_match.group(1).upper()}"
            
            # Check for score data
            scores_match = self.data_pattern.search(line)
            if scores_match and current_group:
                scores = [int(scores_match.group(i)) for i in range(1, 4)]
                
                # Validate scores (should be between 0-100)
                if all(0 <= score <= 100 for score in scores):
                    record = {
                        'date': date,
                        'farm': farm_name,
                        'group': current_group,
                        'score1': scores[0],
                        'score2': scores[1],
                        'score3': scores[2],
                        'total': sum(scores),
                        'average': round(sum(scores) / 3, 1)
                    }
                    extracted_data.append(record)
        
        return extracted_data
    
    def extract_farm_name(self, filename, text):
        """Extract farm name from filename or text"""
        # Try to extract from filename first
        if "sunnyside" in filename.lower():
            return "Sunnyside Farm"
        
        # Look in text
        if "sunnyside" in text.lower():
            return "Sunnyside Farm"
        
        # Default
        return "Unknown Farm"
    
    def extract_date(self, text):
        """Extract date from text"""
        date_match = self.date_pattern.search(text)
        if date_match:
            month, day, year = date_match.groups()
            # Handle 2-digit year
            if len(year) == 2:
                year = "20" + year
            return f"{year}-{month.zfill(2)}-{day.zfill(2)}"
        
        # Default to today's date
        return datetime.now().strftime("%Y-%m-%d")
    
    def process_file(self, file_path):
        """Process a single file and extract data"""
        file_path = Path(file_path)
        
        if file_path.suffix.lower() not in self.supported_formats:
            raise ValueError(f"Unsupported file format: {file_path.suffix}")
        
        all_data = []
        
        if file_path.suffix.lower() == '.pdf':
            # Process PDF
            images = self.pdf_to_images(str(file_path))
            for img in images:
                text = self.extract_text_from_image(img)
                data = self.parse_ocr_text(text, file_path.name)
                all_data.extend(data)
        else:
            # Process image
            processed_img = self.preprocess_image(file_path)
            text = self.extract_text_from_image(processed_img)
            data = self.parse_ocr_text(text, file_path.name)
            all_data.extend(data)
        
        return all_data
    
    def process_folder(self, folder_path):
        """Process all supported files in a folder"""
        folder_path = Path(folder_path)
        all_data = []
        
        for file_path in folder_path.iterdir():
            if file_path.suffix.lower() in self.supported_formats:
                try:
                    logger.info(f"Processing {file_path.name}")
                    data = self.process_file(file_path)
                    all_data.extend(data)
                except Exception as e:
                    logger.error(f"Error processing {file_path.name}: {str(e)}")
        
        return all_data


class DataExporter:
    """Handle data export to various formats"""
    
    @staticmethod
    def to_excel(data, output_path):
        """Export data to Excel with formatting"""
        # Create DataFrame
        df = pd.DataFrame(data)
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Udder Hygiene Data"
        
        # Add headers with formatting
        headers = ['Date', 'Farm Name', 'Group', 'Score 1', 'Score 2', 'Score 3', 'Total', 'Average']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # Add data
        for row_idx, record in enumerate(data, 2):
            ws.cell(row=row_idx, column=1, value=record['date'])
            ws.cell(row=row_idx, column=2, value=record['farm'])
            ws.cell(row=row_idx, column=3, value=record['group'])
            ws.cell(row=row_idx, column=4, value=record['score1'])
            ws.cell(row=row_idx, column=5, value=record['score2'])
            ws.cell(row=row_idx, column=6, value=record['score3'])
            ws.cell(row=row_idx, column=7, value=record['total'])
            ws.cell(row=row_idx, column=8, value=record['average'])
        
        # Add summary statistics
        ws.cell(row=len(data)+4, column=1, value="Summary Statistics")
        ws.cell(row=len(data)+5, column=1, value="Average Score:")
        ws.cell(row=len(data)+5, column=2, value=f"=AVERAGE(H2:H{len(data)+1})")
        
        # Create chart
        chart = BarChart()
        chart.title = "Average Scores by Group"
        chart.y_axis.title = "Average Score"
        chart.x_axis.title = "Group"
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = max(len(str(cell.value or '')) for cell in column)
            ws.column_dimensions[column[0].column_letter].width = max_length + 2
        
        # Save workbook
        wb.save(output_path)
        logger.info(f"Excel file saved to {output_path}")
    
    @staticmethod
    def to_csv(data, output_path):
        """Export data to CSV"""
        df = pd.DataFrame(data)
        df.to_csv(output_path, index=False)
        logger.info(f"CSV file saved to {output_path}")
    
    @staticmethod
    def to_json(data, output_path):
        """Export data to JSON"""
        with open(output_path, 'w') as f:
            json.dump(data, f, indent=2)
        logger.info(f"JSON file saved to {output_path}")


# Flask Web Application
app = Flask(__name__)
app.config['UPLOAD_FOLDER'] = 'uploads'
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB max file size

# Enable CORS for web interface
from flask_cors import CORS
CORS(app)

# Create upload folder if it doesn't exist
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)

ocr_processor = UdderHygieneOCR()
exporter = DataExporter()

@app.route('/')
def home():
    """Welcome route with API documentation"""
    return '''
    <h1>Udder Hygiene OCR API</h1>
    <p>Welcome to the Udder Hygiene Data Automation API!</p>
    <h2>Available Endpoints:</h2>
    <ul>
        <li><strong>GET /api/demo</strong> - Get demo data</li>
        <li><strong>POST /api/upload</strong> - Upload files for OCR processing</li>
        <li><strong>POST /api/export/excel</strong> - Export data to Excel</li>
        <li><strong>POST /api/export/csv</strong> - Export data to CSV</li>
        <li><strong>POST /api/analyze</strong> - Analyze data and get statistics</li>
    </ul>
    <h3>Test the API:</h3>
    <p><a href="/api/demo">Click here to see demo data</a></p>
    '''

@app.route('/api/upload', methods=['POST'])
def upload_files():
    """Handle file upload and OCR processing"""
    if 'files' not in request.files:
        return jsonify({'error': 'No files provided'}), 400
    
    files = request.files.getlist('files')
    all_data = []
    
    for file in files:
        if file.filename == '':
            continue
        
        filename = secure_filename(file.filename)
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        file.save(filepath)
        
        try:
            # Process file with OCR
            data = ocr_processor.process_file(filepath)
            # Ensure data format matches what the frontend expects
            for record in data:
                # Add scores array for compatibility
                if 'scores' not in record:
                    record['scores'] = [record.get('score1', 0), record.get('score2', 0), record.get('score3', 0)]
            all_data.extend(data)
        except Exception as e:
            logger.error(f"Error processing {filename}: {str(e)}")
            return jsonify({'error': f'Error processing {filename}: {str(e)}'}), 500
        finally:
            # Clean up uploaded file
            if os.path.exists(filepath):
                os.remove(filepath)
    
    return jsonify({
        'success': True,
        'data': all_data,
        'count': len(all_data)
    })

@app.route('/api/export/<format>', methods=['POST'])
def export_data(format):
    """Export data to specified format"""
    data = request.json.get('data', [])
    
    if not data:
        return jsonify({'error': 'No data provided'}), 400
    
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    
    try:
        if format == 'excel':
            filename = f"udder_hygiene_{timestamp}.xlsx"
            filepath = os.path.join('exports', filename)
            os.makedirs('exports', exist_ok=True)
            exporter.to_excel(data, filepath)
            return send_file(filepath, as_attachment=True)
            
        elif format == 'csv':
            filename = f"udder_hygiene_{timestamp}.csv"
            filepath = os.path.join('exports', filename)
            os.makedirs('exports', exist_ok=True)
            exporter.to_csv(data, filepath)
            return send_file(filepath, as_attachment=True)
            
        elif format == 'json':
            return jsonify(data)
            
        else:
            return jsonify({'error': 'Invalid format'}), 400
            
    except Exception as e:
        logger.error(f"Export error: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/analyze', methods=['POST'])
def analyze_data():
    """Analyze data and return statistics"""
    data = request.json.get('data', [])
    
    if not data:
        return jsonify({'error': 'No data provided'}), 400
    
    df = pd.DataFrame(data)
    
    # Calculate statistics
    stats = {
        'total_records': len(df),
        'average_score': round(df['average'].mean(), 2),
        'by_group': df.groupby('group')['average'].agg(['mean', 'min', 'max', 'count']).to_dict('index'),
        'by_date': df.groupby('date')['average'].mean().to_dict(),
        'score_distribution': {
            'excellent': len(df[df['average'] >= 90]),
            'good': len(df[(df['average'] >= 80) & (df['average'] < 90)]),
            'fair': len(df[(df['average'] >= 70) & (df['average'] < 80)]),
            'poor': len(df[df['average'] < 70])
        }
    }
    
    return jsonify(stats)

@app.route('/api/demo', methods=['GET'])
def get_demo_data():
    """Return demo data for testing"""
    demo_data = [
        {
            'date': '2025-03-26',
            'farm': 'Sunnyside Farm',
            'group': 'Group A',
            'score1': 85,
            'score2': 92,
            'score3': 88,
            'scores': [85, 92, 88],  # Added for web interface compatibility
            'total': 265,
            'average': 88.3
        },
        {
            'date': '2025-03-26',
            'farm': 'Sunnyside Farm',
            'group': 'Group B',
            'score1': 78,
            'score2': 81,
            'score3': 79,
            'scores': [78, 81, 79],
            'total': 238,
            'average': 79.3
        },
        {
            'date': '2025-03-26',
            'farm': 'Sunnyside Farm',
            'group': 'Group C',
            'score1': 91,
            'score2': 89,
            'score3': 93,
            'scores': [91, 89, 93],
            'total': 273,
            'average': 91.0
        }
    ]
    return jsonify({'success': True, 'data': demo_data})

@app.route('/api/test-ocr', methods=['POST'])
def test_ocr_debug():
    """Debug endpoint to test OCR"""
    try:
        if 'file' not in request.files:
            # Create a simple test image
            import numpy as np
            test_img = np.ones((200, 400, 3), dtype=np.uint8) * 255
            cv2.putText(test_img, "Test 123", (50, 100), cv2.FONT_HERSHEY_SIMPLEX, 2, (0, 0, 0), 3)
            
            # Test OCR
            text = pytesseract.image_to_string(test_img)
            
            return jsonify({
                'success': True,
                'test_image_text': text,
                'tesseract_version': pytesseract.get_tesseract_version().decode('utf-8'),
                'status': 'OCR is working'
            })
        
        # Process uploaded file
        file = request.files['file']
        filename = secure_filename(file.filename)
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        file.save(filepath)
        
        # Read image
        image = cv2.imread(filepath)
        
        # Get raw OCR text
        raw_text = pytesseract.image_to_string(image)
        
        # Also try with preprocessing
        processed_image = ocr_processor.preprocess_image(image)
        processed_text = pytesseract.image_to_string(processed_image)
        
        # Clean up
        os.remove(filepath)
        
        return jsonify({
            'success': True,
            'filename': filename,
            'raw_ocr_text': raw_text[:1000],  # First 1000 chars
            'processed_ocr_text': processed_text[:1000],
            'raw_text_length': len(raw_text),
            'processed_text_length': len(processed_text),
            'tesseract_version': pytesseract.get_tesseract_version().decode('utf-8')
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e),
            'error_type': type(e).__name__
        })
import os

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    app.run(host='0.0.0.0', port=port, debug=False)